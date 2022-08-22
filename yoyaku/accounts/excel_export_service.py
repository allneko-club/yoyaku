import openpyxl as px

from django.utils.timezone import localtime

from yoyaku.accounts.models import User
from yoyaku.booking.models import Booking


def get_customers_xl():
    """顧客一覧のエクセルファイルを作成し返す"""
    wb = px.Workbook()
    ws = wb.active
    ws.title = '顧客データ'

    customers = User.customers.all()

    # 1行目の項目名を出力
    header = ['顧客id', '名前', 'フリガナ', 'メールアドレス', '電話番号', 'LINE名', '年齢', '年代',
              '職業', '郵便番号', '住所', '作業可能時間', '副業経験', 'お問い合わせ内容', 'メモ1', 'メモ2', '削除フラグ', '作成日時']
    ws.append(header)

    #  顧客情報を書き出す
    output_fields = ['id', 'username', 'furigana', 'email', 'phone_number', 'linename', 'age', 'ages', 'job',
                     'zip_code', 'zip', 'workable_time', 'side_business_experience', 'contact', 'memo1', 'memo2', 'is_active', 'date_joined']
    for r, customer in enumerate(customers, 2):
        for c, field in enumerate(output_fields, 1):
            value = getattr(customer, field)
            if field == 'ages':
                value = str(value) if value else ''
            if field == 'date_joined':
                localized_dt = localtime(value)
                value = localized_dt.strftime('%Y-%m-%d %H:%M')
            ws.cell(row=r, column=c).value = value

    # 2列目に列を挿入し、顧客の予約があればスタッフ名を2列目に書き出す
    STAFF_COL = 2
    ws.insert_cols(STAFF_COL)
    ws.cell(row=1, column=STAFF_COL).value = 'スタッフ'
    for r, customer in enumerate(customers, 2):
        booking = Booking.objects.filter(customer=customer.id)
        if len(booking) == 1:
            ws.cell(row=r, column=2).value = booking[0].get_staff_name()

    return wb
